import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import io, os, re
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Reporting Exergon",
    page_icon="⚛️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── PROTECTION PAR MOT DE PASSE ────────────────────────────────────────────────
# Ne s'active que si un mot de passe a été configuré dans Streamlit Cloud
# (Gérer l'application → Settings → Secrets → app_password = "...").
# Si aucun mot de passe n'est configuré, l'app fonctionne normalement sans
# protection (comportement actuel inchangé tant que rien n'est ajouté).
if "app_password" in st.secrets:
    def check_password():
        def password_entered():
            if st.session_state["password"] == st.secrets["app_password"]:
                st.session_state["password_correct"] = True
                del st.session_state["password"]
            else:
                st.session_state["password_correct"] = False

        if st.session_state.get("password_correct", False):
            return True

        st.text_input("Mot de passe", type="password",
                       on_change=password_entered, key="password")
        if "password_correct" in st.session_state and not st.session_state["password_correct"]:
            st.error("Mot de passe incorrect")
        return False

    if not check_password():
        st.stop()

# ─── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
    html, body, [class*="css"], .stApp {
        font-family: 'Space Grotesk', sans-serif;
        background-color: #ffffff !important;
        color: #1a2e1a !important;
    }
    .stApp { background: #ffffff !important; }
    .block-container { background: transparent !important; padding-top: 1.5rem !important; }

    [data-testid="stSidebar"] {
        background: #f4f9f4 !important;
        border-right: 1px solid #c8e6c8 !important;
    }
    [data-testid="stSidebar"] * { color: #1a3a1a !important; }
    [data-testid="stSidebar"] label { color: #1a5c1a !important; font-size:0.78rem; }

    .main-header {
        background: linear-gradient(135deg, #0a2e0a 0%, #0f3d0f 50%, #143d14 100%);
        border-left: 4px solid #2d8a2d;
        padding: 2rem 2.5rem; border-radius: 16px; margin-bottom: 2rem;
        position: relative; overflow: hidden;
    }
    .main-header::before {
        content: "";
        position: absolute; top: -60px; right: -60px;
        width: 200px; height: 200px;
        background: radial-gradient(circle, rgba(45,138,45,0.15) 0%, transparent 70%);
        border-radius: 50%;
    }
    .main-header h1 { font-size: 1.9rem; font-weight: 700; margin: 0; color: #ffffff; letter-spacing: -0.5px; }
    .main-header h1 span { color: #6fcf97; }
    .main-header p { margin: 0.3rem 0 0; color: #a7d7a7; font-size: 0.85rem; }
    .header-badge {
        display: inline-block; background: rgba(45,138,45,0.2);
        border: 1px solid #2d8a2d; color: #6fcf97;
        font-size: 0.7rem; font-weight: 700; letter-spacing: 1px;
        padding: 0.2rem 0.6rem; border-radius: 4px; margin-bottom: 0.6rem;
        font-family: 'JetBrains Mono', monospace;
    }

    .kpi-card {
        background: #ffffff; border: 1px solid #c8e6c8; border-radius: 14px;
        padding: 1.4rem 1.6rem;
        box-shadow: 0 2px 12px rgba(0,80,0,0.06);
        transition: transform 0.15s, box-shadow 0.15s, border-color 0.15s;
        height: 100%; position: relative; overflow: hidden;
    }
    .kpi-card::after {
        content: ""; position: absolute; bottom: 0; left: 0; right: 0; height: 2px;
        background: linear-gradient(90deg, transparent, #2d8a2d, transparent);
        opacity: 0; transition: opacity 0.2s;
    }
    .kpi-card:hover { transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,80,0,0.12); border-color: #2d8a2d; }
    .kpi-card:hover::after { opacity: 1; }
    .kpi-label {
        font-size: 0.72rem; font-weight: 600; color: #4a7a4a;
        text-transform: uppercase; letter-spacing: 1px; margin-bottom: 0.5rem;
        font-family: 'JetBrains Mono', monospace;
    }
    .kpi-value { font-size: 1.85rem; font-weight: 700; color: #0a3d0a; line-height: 1; }
    .kpi-sub   { font-size: 0.78rem; color: #4a7a4a; margin-top: 0.45rem; }
    .kpi-badge {
        display: inline-block; font-size: 0.7rem; font-weight: 600;
        padding: 0.15rem 0.6rem; border-radius: 20px; margin-top: 0.45rem;
        font-family: 'JetBrains Mono', monospace; letter-spacing: 0.3px;
    }
    .badge-green  { background: #ecfdf5; color: #065f46; border: 1px solid #a7f3d0; }
    .badge-orange { background: #fffbeb; color: #92400e; border: 1px solid #fcd34d; }
    .badge-red    { background: #fef2f2; color: #991b1b; border: 1px solid #fca5a5; }
    .badge-blue   { background: #eff6ff; color: #1e40af; border: 1px solid #bfdbfe; }

    .section-title {
        font-size: 0.72rem; font-weight: 600; color: #1a5c1a;
        margin: 2rem 0 1rem; padding-bottom: 0.5rem;
        border-bottom: 1px solid #c8e6c8;
        text-transform: uppercase; letter-spacing: 2px;
        font-family: 'JetBrains Mono', monospace;
    }

    [data-testid="stTabs"] button { color: #4a7a4a !important; font-weight: 600 !important; font-size: 0.85rem !important; }
    [data-testid="stTabs"] button[aria-selected="true"] { color: #0a3d0a !important; border-bottom-color: #2d8a2d !important; }
    [data-testid="stTabs"] { border-bottom: 1px solid #c8e6c8 !important; }

    .stTextInput input { background: #f4f9f4 !important; border: 1px solid #c8e6c8 !important; color: #1a2e1a !important; border-radius: 8px !important; }
    .stTextInput input:focus { border-color: #2d8a2d !important; }
    .stCaption { color: #4a7a4a !important; font-family: 'JetBrains Mono', monospace; font-size: 0.72rem; }

    .stDownloadButton button {
        background: #ecfdf5 !important; border: 1px solid #2d8a2d !important;
        color: #065f46 !important; font-weight: 600 !important; border-radius: 8px !important; transition: all 0.2s !important;
    }
    .stDownloadButton button:hover { background: #d1fae5 !important; box-shadow: 0 4px 12px rgba(0,80,0,0.15) !important; }

    .legend-box { display:flex; gap:1.5rem; margin-bottom:0.8rem; flex-wrap:wrap; }
    .legend-item { display:flex; align-items:center; gap:0.4rem; font-size:0.78rem; color:#4a7a4a; }
    .legend-dot  { width:12px; height:12px; border-radius:3px; }
    .dot-green-dark  { background:#1a6e3c; }
    .dot-green-light { background:#6fcf97; border: 1px solid #a7f3d0; }

    #MainMenu, footer { visibility: hidden; }
    .stMultiSelect [data-baseweb="tag"] { background: #d1fae5 !important; }
</style>
""", unsafe_allow_html=True)

# ---- CHARGEMENT DATA ----------------------------------------------------
def clean_html(text):
    if not text or str(text).lower() in ["nan","none",""]:
        return ""
    text = re.sub(r'<[^>]+>', '', str(text))
    return text.replace("&nbsp;", " ").replace("&amp;", "&").strip()

@st.cache_data
def load_data(path):
    return load_file(path)

def load_file(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, header=2)

    # --- Normalisation du nom de colonne "Étape" (certains fichiers l'exportent
    # sans accent : "Etape"). On uniformise pour que le reste du code (filtres,
    # KPIs, graphiques) qui utilise "Étape" fonctionne toujours.
    if "Étape" not in df.columns and "Etape" in df.columns:
        df = df.rename(columns={"Etape": "Étape"})

    # --- Revenu : ancien format "Revenu attendu" ou nouveau format "Ticket"
    if "Revenu attendu" in df.columns:
        df["Revenu_M"] = pd.to_numeric(df["Revenu attendu"], errors="coerce").fillna(0) / 1_000_000
        if "Ticket" not in df.columns:
            df["Ticket"] = df["Revenu_M"].apply(lambda x: f"{x:.0f} M€")
    elif "Ticket" in df.columns:
        raw_ticket = df["Ticket"]
        # NB: on utilise pd.isna() plutôt qu'un simple .astype(str) vectorisé,
        # car avec les dtypes "str" de pandas récents, une valeur NaN convertie
        # en chaîne ne redevient pas forcément la string "nan" (d'où l'ancien bug).
        ticket_vide = raw_ticket.apply(
            lambda x: pd.isna(x) or str(x).strip().lower() in ["nan", "none", "", "—"]
        )

        # Colonne cachée sans nom (souvent en toute fin de fichier) qui contient
        # le vrai montant en euros quand la colonne "Ticket" texte est vide.
        last_col = df.columns[-1]
        if ticket_vide.any() and str(last_col).startswith("Unnamed"):
            montant_cache = pd.to_numeric(df[last_col], errors="coerce")
        else:
            montant_cache = pd.Series([None] * len(df), index=df.index)

        def format_ticket(is_vide, val_texte, montant):
            if not is_vide:
                return str(val_texte)
            if pd.notna(montant):
                return f"{montant/1_000_000:.0f} M€"
            return "—"

        df["Ticket"] = [
            format_ticket(v, t, m)
            for v, t, m in zip(ticket_vide, raw_ticket, montant_cache)
        ]

        # Calcul numérique pour les KPIs
        def parse_ticket(val):
            v = str(val).strip().upper()
            try:
                n = float(v.replace("M€","").replace("K€","").replace(" ","").replace(",","."))
                return n / 1000 if "K€" in v else n
            except:
                return 0.0
        df["Revenu_M"] = df["Ticket"].apply(parse_ticket)
    else:
        df["Revenu_M"] = 0.0
        df["Ticket"] = "—"

    # ── Maturation : ancien "Matu." ou nouveau "% Maturation"
    matu_col = "Matu." if "Matu." in df.columns else "% Maturation" if "% Maturation" in df.columns else None
    if matu_col:
        df["Matu_num"] = pd.to_numeric(
            df[matu_col].astype(str).str.replace("%","",regex=False).str.strip(), errors="coerce"
        ).fillna(0)
        df["Matu."] = df[matu_col]
    else:
        df["Matu_num"] = 0.0
        df["Matu."] = ""

    df["Pipeline_pondere"] = df["Revenu_M"] * df["Matu_num"] / 100

    # ── Dates : détection automatique du nom de colonne
    col_d = next((c for c in df.columns if "derni" in c.lower() and "activit" in c.lower() and "date" in c.lower()), None)
    col_p = next((c for c in df.columns if "proch" in c.lower() and "activit" in c.lower() and "date" in c.lower()), None)

    if col_d:
        df["_date_derniere"] = pd.to_datetime(
            df[col_d].astype(str).str.replace("-","",regex=False).str.strip(),
            dayfirst=True, errors="coerce"
        )
        df["Date de dernière activité"] = df[col_d]
    else:
        df["_date_derniere"] = pd.NaT
        df["Date de dernière activité"] = ""

    if col_p:
        df["_date_prochaine"] = pd.to_datetime(
            df[col_p].astype(str).str.replace("-","",regex=False).str.strip(),
            dayfirst=True, errors="coerce"
        )
        df["Date de prochaine activité"] = df[col_p]
    else:
        df["_date_prochaine"] = pd.NaT
        df["Date de prochaine activité"] = ""

    # ── Dernier commentaire : ancien multi-colonnes ou nouveau colonne unique
    def last_comment(row):
        for col in ["Commentaire4","Commentaire3","Commentaire2","Commentaire"]:
            v = clean_html(row.get(col,""))
            if v: return v[:400]
        return ""
    df["Dernier commentaire"] = df.apply(last_comment, axis=1)

    # ── Colonnes manquantes — on les crée vides pour éviter les KeyError
    for col in ["Étape","Intérêt","Tri","Fonds","Typologie","Entité","Contact","Contact Levée","Ticket"]:
        if col not in df.columns:
            df[col] = ""

    return df

@st.cache_data
def load_data(path: str) -> pd.DataFrame:
    return load_file(path)

def load_data_no_cache(path: str) -> pd.DataFrame:
    return load_file(path)

# ─── SIDEBAR ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div style="font-family:JetBrains Mono,monospace;font-size:1rem;font-weight:700;color:#39ff14;letter-spacing:2px;padding:0.5rem 0 0.2rem">⚛ EXERGON</div><div style="font-size:0.65rem;color:#4a8a4a;letter-spacing:1.5px;margin-bottom:0.5rem;font-family:JetBrains Mono,monospace">NUCLEAR ENERGY FUND</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("**📂 Fichier principal**")
    uploaded_main = st.file_uploader("Glisse ton fichier xlsx", type=["xlsx"], key="main")
    st.markdown("**📅 Fichiers historiques** *(optionnel)*")
    st.caption("Ajoute plusieurs fichiers xlsx pour voir l'évolution dans le temps.")
    uploaded_history = st.file_uploader(
        "Fichiers passés", type=["xlsx"], accept_multiple_files=True, key="history"
    )
    history_labels = []
    if uploaded_history:
        st.caption("Date de chaque fichier (JJ/MM/AAAA) :")
        for i, f in enumerate(uploaded_history):
            label = st.text_input(f"Date — {f.name}", key=f"lbl_{i}", placeholder="01/05/2026")
            history_labels.append(label if label else f.name)
    st.markdown("---")
    st.markdown("**Filtres**")

# ─── CHARGEMENT DEPUIS DOSSIER HISTORIQUE ──────────────────────────────────────
def _parse_date_from_filename(name):
    m = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", name)
    if m:
        return pd.Timestamp(f"{m.group(1)}-{m.group(2)}-{m.group(3)}")
    return None

def _get_historique_files_local():
    """Ancien comportement : lit le dossier historique/ présent dans le dépôt
    lui-même (utilisé si aucun dépôt de données privé n'est configuré)."""
    hist_dir = "historique"
    if not os.path.exists(hist_dir):
        return []
    files = []
    for f in os.listdir(hist_dir):
        if f.endswith(".xlsx") and not f.startswith("~"):
            path = os.path.join(hist_dir, f)
            date_ts = _parse_date_from_filename(f) or pd.Timestamp(os.path.getmtime(path), unit="s")
            files.append({"path": path, "name": f, "date": date_ts,
                          "label": date_ts.strftime("%d/%m/%Y")})
    files.sort(key=lambda x: x["date"])
    return files

@st.cache_data(ttl=300, show_spinner="Chargement des fichiers historiques…")
def _get_historique_files_from_private_repo(data_repo, data_path, _token):
    """Récupère les fichiers xlsx depuis un dépôt GitHub PRIVÉ (data_repo =
    'Organisation/nom-du-depot'), dans le sous-dossier data_path (utile pour un
    dépôt centralisé avec un dossier par fonds, ex. 'calderion/historique').
    Nécessite les secrets Streamlit `github_token`, `data_repo` et `data_path`."""
    headers = {"Authorization": f"token {_token}", "Accept": "application/vnd.github+json"}
    api_url = f"https://api.github.com/repos/{data_repo}/contents/{data_path}"
    resp = requests.get(api_url, headers=headers, timeout=15)
    if resp.status_code != 200:
        st.error(
            f"Impossible de lire '{data_path}' dans le dépôt privé '{data_repo}' "
            f"(code {resp.status_code}). Vérifie les secrets 'github_token', "
            f"'data_repo' et 'data_path' dans Streamlit Cloud."
        )
        return []

    os.makedirs("/tmp/historique_cache", exist_ok=True)
    files = []
    for item in resp.json():
        name = item.get("name", "")
        if not name.endswith(".xlsx") or name.startswith("~"):
            continue
        raw_resp = requests.get(
            item["url"],
            headers={**headers, "Accept": "application/vnd.github.raw"},
            timeout=15,
        )
        if raw_resp.status_code != 200:
            continue
        local_path = f"/tmp/historique_cache/{name}"
        with open(local_path, "wb") as f:
            f.write(raw_resp.content)
        date_ts = _parse_date_from_filename(name) or pd.Timestamp.now()
        files.append({"path": local_path, "name": name, "date": date_ts,
                      "label": date_ts.strftime("%d/%m/%Y")})
    files.sort(key=lambda x: x["date"])
    return files

def get_historique_files():
    """Bascule automatiquement : si les secrets 'github_token' et 'data_repo'
    sont configurés, va chercher les fichiers dans le dépôt privé (dans le
    sous-dossier 'data_path', défaut 'historique' — pratique pour un dépôt
    centralisé avec un dossier par fonds) ; sinon, comportement historique
    (dossier historique/ local au dépôt public)."""
    data_repo = st.secrets.get("data_repo")
    token = st.secrets.get("github_token")
    data_path = st.secrets.get("data_path", "historique")
    if data_repo and token:
        return _get_historique_files_from_private_repo(data_repo, data_path, token)
    return _get_historique_files_local()

hist_files = get_historique_files()

# Debug temporaire — à supprimer après
import os as _os
_hist_dir = "historique"
_exists = _os.path.exists(_hist_dir)
_cwd = _os.getcwd()
_files_in_hist = _os.listdir(_hist_dir) if _exists else []
_all_files = _os.listdir(".")
with st.sidebar:
    with st.expander("🔍 Debug (temporaire)"):
        st.write(f"CWD: {_cwd}")
        st.write(f"Source historique: {'dépôt privé (' + st.secrets.get('data_repo','') + ')' if st.secrets.get('data_repo') else 'dossier local'}")
        st.write(f"Dossier historique local existe: {_exists}")
        st.write(f"Fichiers dans historique/ local: {_files_in_hist}")
        st.write(f"Fichiers racine: {_all_files}")
        st.write(f"hist_files trouvés: {[h['name'] for h in hist_files]}")

# Fichier le plus récent = fichier principal
if uploaded_main:
    tmp_path = "/tmp/data_main.xlsx"
    with open(tmp_path, "wb") as f:
        f.write(uploaded_main.getvalue())
    df_raw = load_data(tmp_path)
    current_label = "Upload manuel"
    current_date  = pd.Timestamp.now()
elif hist_files:
    latest = hist_files[-1]
    df_raw = load_data(latest["path"])
    current_label = latest["label"]
    current_date  = latest["date"]
    # Fichiers historiques = tous sauf le plus récent, max 4 derniers
    hist_files_display = hist_files[:-1][-4:]
else:
    st.info("👋 Bienvenue ! Dépose ton fichier Excel dans le dossier **historique/** sur GitHub pour commencer.")
    st.stop()

if "hist_files_display" not in dir():
    hist_files_display = []

# ─── FILTRES ────────────────────────────────────────────────────────────────────
with st.sidebar:
    sel_typo   = st.multiselect("Typologie",   sorted(df_raw["Typologie"].dropna().unique()),   default=sorted(df_raw["Typologie"].dropna().unique()))
    sel_etape  = st.multiselect("Étape",       sorted(df_raw["Étape"].dropna().unique()),       default=sorted(df_raw["Étape"].dropna().unique()))
    sel_interet= st.multiselect("Intérêt",     sorted(df_raw["Intérêt"].dropna().unique()),     default=sorted(df_raw["Intérêt"].dropna().unique()))
    rev_min, rev_max = float(df_raw["Revenu_M"].min()), float(df_raw["Revenu_M"].max())
    rev_range = st.slider("Revenu (M€)", rev_min, max(rev_max, rev_min+0.01), (rev_min, max(rev_max, rev_min+0.01)), step=0.5) if rev_max > rev_min else (rev_min, rev_max)
    sel_tri    = st.multiselect("Tri activités", sorted(df_raw["Tri"].dropna().unique()), default=sorted(df_raw["Tri"].dropna().unique()))

df = df_raw.copy()
if sel_typo:    df = df[df["Typologie"].isin(sel_typo)]
if sel_etape:   df = df[df["Étape"].isin(sel_etape)]
if sel_interet: df = df[df["Intérêt"].isin(sel_interet)]
if sel_tri:     df = df[df["Tri"].isin(sel_tri)]
df = df[df["Revenu_M"].between(rev_range[0], rev_range[1])]

# ─── HEADER ─────────────────────────────────────────────────────────────────────
today = datetime.today()
st.markdown(f"""
<div class="main-header">
    <div class="header-badge">⚛️ NUCLEAR ENERGY FUND</div>
    <h1>EXERGON — <span>Pipeline Dashboard</span></h1>
    <p>Données au {current_label} &nbsp;·&nbsp; {len(df)} leads filtrés sur {len(df_raw)}</p>
</div>
""", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["📈  KPIs & Analyse", "📋  Pipeline Détaillé"])

# ════════════════════════════════════════════════════════════════════════════════
# TAB 1 — KPIs
# ════════════════════════════════════════════════════════════════════════════════
with tab1:
    total_leads = len(df)
    ok_count    = (df["Intérêt"]=="OK").sum()
    ko_count    = (df["Intérêt"]=="KO").sum()
    conv_rate   = ok_count/total_leads*100 if total_leads else 0
    en_retard   = df[df["_date_prochaine"].notna() & (df["_date_prochaine"]<pd.Timestamp.now())].shape[0]
    semaine_passee = df["Tri"].isin(["Semaine passée","Semaine passée & à venir"]).sum()
    semaine_venir  = df["Tri"].isin(["À venir","Semaine passée & à venir"]).sum()
    df_ok_typo = df[df["Intérêt"] == "OK"]
    top_typo = df_ok_typo.groupby("Typologie")["Revenu_M"].sum().idxmax() if len(df_ok_typo) else "—"
    top_val  = df_ok_typo.groupby("Typologie")["Revenu_M"].sum().max()    if len(df_ok_typo) else 0

    st.markdown('<div class="section-title">Vue d\'ensemble</div>', unsafe_allow_html=True)
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    with c1:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">👥 Nombre de Leads</div>
            <div class="kpi-value">{total_leads}</div>
            <div class="kpi-sub">au {current_label}</div></div>""", unsafe_allow_html=True)
    with c2:
        br="badge-red" if en_retard>10 else "badge-orange" if en_retard>5 else "badge-green"
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">⏰ Activités en Retard</div>
            <div class="kpi-value">{en_retard}</div>
            <span class="kpi-badge {br}">relances à planifier</span></div>""", unsafe_allow_html=True)
    with c3:
        bc="badge-green" if conv_rate>=30 else "badge-orange" if conv_rate>=10 else "badge-red"
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">✅ Taux de Conversion</div>
            <div class="kpi-value">{conv_rate:.1f}%</div>
            <span class="kpi-badge {bc}">{ok_count} OK · {ko_count} KO</span></div>""", unsafe_allow_html=True)
    with c4:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">📅 Activités Semaine Passée</div>
            <div class="kpi-value">{semaine_passee}</div>
            <div class="kpi-sub">activités réalisées</div></div>""", unsafe_allow_html=True)
    with c5:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">🔜 Activités à Venir</div>
            <div class="kpi-value">{semaine_venir}</div>
            <div class="kpi-sub">dans les 2 prochaines semaines</div></div>""", unsafe_allow_html=True)
    with c6:
        st.markdown(f"""<div class="kpi-card"><div class="kpi-label">🏆 Top Typologie</div>
            <div class="kpi-value" style="font-size:1.1rem;padding-top:.3rem">{top_typo}</div>
            <span class="kpi-badge badge-blue">{top_val:.1f} M€</span></div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Graphiques évolutifs (empilés par période + courbe de total) ────────────
    all_snapshots_available = hist_files_display or uploaded_history
    if all_snapshots_available:
        st.markdown('<div class="section-title">📅 Évolution dans le Temps</div>', unsafe_allow_html=True)

        snapshots = []

        # Fichiers depuis dossier historique/
        for h in hist_files_display:
            dh = load_file(h["path"])
            snapshots.append({"df": dh, "label": h["label"], "ts": h["date"]})

        # Fichiers uploadés manuellement (optionnel)
        if uploaded_history:
            for i, hf in enumerate(uploaded_history):
                tmp = f"/tmp/hist_{i}.xlsx"
                with open(tmp,"wb") as f: f.write(hf.getvalue())
                dh = load_file(tmp)
                lbl = history_labels[i] if i < len(history_labels) else hf.name
                ts  = pd.to_datetime(lbl, dayfirst=True, errors="coerce")
                snapshots.append({"df": dh, "label": lbl, "ts": ts})

        # Fichier actuel
        snapshots.append({"df": df_raw, "label": current_label, "ts": current_date})
        snapshots.sort(key=lambda x: x["ts"] if pd.notna(x["ts"]) else pd.Timestamp.min)

        # ── Palettes de couleurs ──────────────────────────────────────────────
        # Étape : ordre métier fixe, du vert foncé (Mort) au vert clair (Souscription)
        ETAPE_ORDER = [
            "Mort / Plus rien à faire",
            "Probablement mort",
            "Pas encore prêt - A alimenter",
            "Intérêt à qualifier",
            "Interêt indicatif - A creuser",
            "Due diligence",
            "Souscription réalisée",
        ]
        ETAPE_GREENS = ["#1B4332","#2D6A4F","#40916C","#52B788","#74C69D","#95D5B2","#D8F3DC"]
        ETAPE_COLOR_MAP = dict(zip(ETAPE_ORDER, ETAPE_GREENS))
        GREY_FALLBACK = "#C9C9C9"  # pour "Non renseigné" ou toute étape imprévue

        def blue_palette(n):
            return px.colors.sample_colorscale(
                [[0.0,"#03045E"],[0.3,"#0077B6"],[0.6,"#00B4D8"],[1.0,"#CAF0F8"]],
                [i/max(n-1,1) for i in range(n)]
            )

        LINE_COLOR = "#D6672B"  # accent chaud, contrastant avec le vert et le bleu

        # ── Axe des mois : on réserve la place pour les 12 mois de l'année ─────
        MONTHS_FR = ["Jan","Fév","Mars","Avr","Mai","Juin","Juil","Août","Sept","Oct","Nov","Déc"]

        def month_label(ts):
            if pd.isna(ts):
                return None
            return MONTHS_FR[ts.month - 1]

        def evolution_stacked_chart(group_col, title, value_col=None, matu_range=None,
                                     y_label="Nb Leads", ok_only=False,
                                     color_map=None, dynamic_palette=None,
                                     top_n=None, show_line=False):
            """Construit un graphique en barres empilées par mois (un fichier historique =
            un mois), regroupées par group_col (Étape ou Typologie). Si show_line=True,
            une courbe fine est ajoutée sur un axe secondaire à droite, indiquant le
            nombre de leads (indépendant de l'axe principal en valeur/M€). L'axe des
            mois couvre toute l'année pour garder une taille cohérente au fil des ajouts
            de fichiers."""
            rows, nb_leads_rows = [], []
            for s in snapshots:
                lbl = month_label(s["ts"])
                if lbl is None:
                    continue
                d = s["df"].copy()
                if matu_range:
                    d = d[d["Matu_num"].between(matu_range[0], matu_range[1])]
                if ok_only:
                    d = d[d["Intérêt"] == "OK"]
                d[group_col] = d[group_col].fillna("Non renseigné")
                d.loc[d[group_col].astype(str).str.strip()=="", group_col] = "Non renseigné"
                if value_col:
                    grp = d.groupby(group_col)[value_col].sum()
                else:
                    grp = d.groupby(group_col).size()
                for cat, val in grp.items():
                    if val == 0:
                        continue
                    rows.append({"Période": lbl, group_col: cat, "Valeur": val})
                nb_leads_rows.append({"Période": lbl, "Nb Leads": len(d)})

            df_long = pd.DataFrame(rows)
            df_nb   = pd.DataFrame(nb_leads_rows)

            # Limiter aux N catégories les plus importantes (en nombre cumulé) ; le
            # reste est regroupé dans "Autre" pour que le total des barres reste cohérent.
            if top_n and len(df_long):
                top_cats = df_long.groupby(group_col)["Valeur"].sum().nlargest(top_n).index
                df_long[group_col] = df_long[group_col].where(df_long[group_col].isin(top_cats), "Autre")
                df_long = df_long.groupby(["Période", group_col], as_index=False)["Valeur"].sum()

            # Couleurs : soit une correspondance fixe (Étape), soit un dégradé dynamique
            # du plus foncé (catégorie la plus importante) au plus clair (Typologie)
            if color_map is not None:
                cats_present = df_long[group_col].unique().tolist() if len(df_long) else []
                cmap = {c: color_map.get(c, GREY_FALLBACK) for c in cats_present}
                order = [c for c in color_map.keys() if c in cats_present] + \
                        [c for c in cats_present if c not in color_map]
            elif dynamic_palette is not None and len(df_long):
                order = [c for c in df_long.groupby(group_col)["Valeur"].sum()
                         .sort_values(ascending=False).index.tolist() if c != "Autre"]
                colors = dynamic_palette(len(order))
                cmap = dict(zip(order, colors))
                if "Autre" in df_long[group_col].values:
                    cmap["Autre"] = "#B0B0B0"
                    order = order + ["Autre"]
            else:
                cmap = None
                order = None

            df_long["Texte"] = df_long["Valeur"].apply(
                lambda v: f"{v:.1f}" if value_col else f"{int(v)}"
            )

            fig = px.bar(df_long, x="Période", y="Valeur", color=group_col, barmode="stack",
                         title=title, color_discrete_map=cmap,
                         category_orders={"Période": MONTHS_FR, **({group_col: order} if order else {})},
                         text="Texte",
                         labels={"Valeur": y_label, "Période": ""})
            fig.update_traces(marker_line_width=0, textposition="inside",
                              insidetextanchor="middle", textfont_size=9, textfont_color="white")

            layout_kwargs = dict(
                plot_bgcolor="#f4f9f4", paper_bgcolor="#ffffff",
                font_family="Space Grotesk", font_color="#1a2e1a",
                margin=dict(l=10,r=10,t=50,b=60),
                legend=dict(orientation="h", y=-0.25, font_size=10),
                xaxis=dict(categoryorder="array", categoryarray=MONTHS_FR, range=[-0.5, 11.5]),
                yaxis=dict(title=y_label, title_font_color="#000000",
                           title_standoff=4, tickfont_color="#000000"),
            )

            if show_line:
                fig.add_trace(go.Scatter(
                    x=df_nb["Période"], y=df_nb["Nb Leads"],
                    mode="lines+markers", name="Nb Leads",
                    line=dict(color=LINE_COLOR, width=1.6),
                    marker=dict(size=6, color=LINE_COLOR),
                    yaxis="y2", showlegend=False
                ))
                layout_kwargs["yaxis2"] = dict(
                    title="Nb Leads", overlaying="y", side="right",
                    title_font_color="#000000", title_standoff=4,
                    tickfont_color="#000000", showgrid=False,
                )

            fig.update_layout(**layout_kwargs)
            return fig

        ev1, ev2 = st.columns(2)
        with ev1:
            fig = evolution_stacked_chart("Étape", "La Dynamique — Nombre de Leads par Étape",
                                          color_map=ETAPE_COLOR_MAP, show_line=False)
            st.plotly_chart(fig, use_container_width=True)
        with ev2:
            fig = evolution_stacked_chart("Étape", "Le Résultat — Valeur par Étape (Matu. 30–80%)",
                                          value_col="Revenu_M", matu_range=(30,80), y_label="M€",
                                          color_map=ETAPE_COLOR_MAP, show_line=True)
            st.plotly_chart(fig, use_container_width=True)

        ev3, ev4 = st.columns(2)
        with ev3:
            fig = evolution_stacked_chart("Typologie", "TOP 5 Typologies OK", ok_only=True,
                                          dynamic_palette=blue_palette, top_n=5, show_line=False)
            st.plotly_chart(fig, use_container_width=True)
        with ev4:
            fig = evolution_stacked_chart("Typologie", "Valeur par Typologie (Matu. 30–80%)",
                                          value_col="Revenu_M", matu_range=(30,80), y_label="M€",
                                          dynamic_palette=blue_palette, show_line=True)
            st.plotly_chart(fig, use_container_width=True)

        # ── Tableau comparatif des activités
        st.markdown('<div class="section-title">📊 Comparatif Activités d\'une Période à l\'Autre</div>', unsafe_allow_html=True)
        act_rows = []
        for s in snapshots:
            d = s["df"]
            act_rows.append({
                "Période": s["label"],
                "Nb activités enregistrées": int(d["_date_derniere"].notna().sum()),
                "Leads avec prochaine activité": int(d["_date_prochaine"].notna().sum()),
                "Leads sans activité planifiée": int(d["_date_prochaine"].isna().sum()),
                "Activités en retard": int((d["_date_prochaine"] < pd.Timestamp.now()).sum()),
            })
        df_act = pd.DataFrame(act_rows)

        # Calcul delta vs période précédente
        if len(df_act) > 1:
            for col in df_act.columns[1:]:
                prev = df_act[col].shift(1)
                df_act[f"Δ {col}"] = (df_act[col] - prev).apply(
                    lambda x: f"+{int(x)}" if pd.notna(x) and x > 0 else (f"{int(x)}" if pd.notna(x) else "—")
                )

        st.dataframe(df_act, use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════════════════════════════════════
# TAB 2 — TABLEAU
# ════════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="section-title">Pipeline Détaillé</div>', unsafe_allow_html=True)

    COLS_DISPLAY = [
        "Fonds","Typologie","Entité","Contact","Contact Levée",
        "Dernier commentaire","Intérêt",
        "Date de dernière activité","Date de prochaine activité",
        "Ticket","Matu.","Tri"
    ]

    fc1, fc2, fc3 = st.columns([2,2,3])
    with fc1: search   = st.text_input("🔍 Rechercher", "")
    with fc2: sort_col = st.selectbox("Trier par", ["Entité","Revenu_M","Matu_num","Étape","Typologie"], index=2)
    with fc3: sort_dir = st.radio("Ordre", ["Décroissant","Croissant"], horizontal=True)

    df_disp = df.copy()
    df_disp["Date de dernière activité"]  = df_disp["_date_derniere"].dt.strftime("%d/%m/%Y").fillna("")
    df_disp["Date de prochaine activité"] = df_disp["_date_prochaine"].dt.strftime("%d/%m/%Y").fillna("")

    df_disp = df_disp[[c for c in COLS_DISPLAY if c in df_disp.columns]].copy()
    df_disp = df_disp.rename(columns={
        "Date de dernière activité":  "Dernière activité",
        "Date de prochaine activité": "Prochaine activité"
    })

    if search:
        mask = df_disp.apply(lambda col: col.astype(str).str.contains(search, case=False, na=False)).any(axis=1)
        df_disp = df_disp[mask]

    sort_map = {"Revenu_M":"Revenu_M","Matu_num":"Matu."}
    sort_key = sort_map.get(sort_col, sort_col)
    if sort_key in df_disp.columns:
        df_disp = df_disp.sort_values(sort_key, ascending=(sort_dir=="Croissant"))

    # ── Tableau HTML — wrapping réel + tri JS + couleurs
    now_ts           = pd.Timestamp.now()
    cutoff_derniere  = now_ts - timedelta(days=9)
    cutoff_prochaine = now_ts + timedelta(days=13)

    COLS = [
        ("Fonds",               "110px"),
        ("Typologie",           "110px"),
        ("Entité",              "120px"),
        ("Contact",             "100px"),
        ("Contact Levée",       "100px"),
        ("Dernier commentaire", "420px"),
        ("Intérêt",             "65px"),
        ("Dernière activité",   "100px"),
        ("Prochaine activité",  "100px"),
        ("Ticket",              "65px"),
        ("Matu.",               "55px"),
        ("Tri",                 "120px"),
    ]

    def cell_style(col, flag_d, flag_p, row_bg):
        base = f"padding:9px 10px;vertical-align:middle;border-bottom:1px solid #e2f0e2;line-height:1.5;font-size:0.79rem;background:{row_bg};"
        if col == "Dernier commentaire":
            base = base.replace("vertical-align:middle", "vertical-align:top")
            base += "white-space:normal;word-break:break-word;"
        if flag_d and col in ["Dernière activité","Tri","Dernier commentaire"]:
            return base.replace(f"background:{row_bg};","background:#1a6e3c;color:#fff;")
        if flag_p and col in ["Prochaine activité","Tri","Dernier commentaire"]:
            return base.replace(f"background:{row_bg};","background:#6fcf97;color:#0a3d0a;")
        return base

    # Header
    ths = ""
    for i,(col,w) in enumerate(COLS):
        ths += (f'<th onclick="srt({i})" '
                f'style="width:{w};min-width:{w};padding:10px;cursor:pointer;'
                f'white-space:nowrap;border-right:1px solid #1a5c1a;'
                f'font-size:0.7rem;text-transform:uppercase;letter-spacing:0.8px;">'
                f'{col} <span id="si{i}" style="opacity:.45;font-size:.65rem;">&#8645;</span></th>')

    # Rows
    rows_parts = []
    for idx,(_, row) in enumerate(df_disp.iterrows()):
        date_d = pd.to_datetime(row.get("Dernière activité",""), dayfirst=True, errors="coerce")
        date_p = pd.to_datetime(row.get("Prochaine activité",""), dayfirst=True, errors="coerce")
        flag_d = pd.notna(date_d) and date_d >= cutoff_derniere
        flag_p = pd.notna(date_p) and now_ts <= date_p <= cutoff_prochaine
        bg = "#f4faf4" if idx % 2 == 0 else "#ffffff"
        cells = ""
        for col,_ in COLS:
            val = str(row.get(col,"") or "").replace("<","&lt;").replace(">","&gt;")
            cells += f"<td style='{cell_style(col,flag_d,flag_p,bg)}'>{val}</td>"
        rows_parts.append(f"<tr>{cells}</tr>")

    rows_html = "\n".join(rows_parts)

    table = (
        '<div style="display:flex;justify-content:flex-end;margin-bottom:6px;">'
        '<button onclick="document.getElementById(\'twrap\').requestFullscreen()" '
        'style="background:#0a3d0a;color:#fff;border:none;border-radius:8px;'
        'padding:6px 14px;font-family:Space Grotesk,sans-serif;font-size:0.75rem;'
        'cursor:pointer;display:flex;align-items:center;gap:6px;">'
        '⛶ Plein écran</button></div>'
        '<div id="twrap" style="overflow:auto;max-height:650px;border-radius:12px;'
        'border:1px solid #c8e6c8;font-family:Space Grotesk,sans-serif;">'
        '<table id="pt" style="border-collapse:collapse;width:100%;min-width:1400px;">'
        '<thead><tr id="hdr" style="position:sticky;top:0;z-index:10;'
        'background:#0a3d0a;color:#fff;">'
        + ths +
        '</tr></thead>'
        '<tbody id="ptb">' + rows_html + '</tbody>'
        '</table></div>'
        '<style>'
        '#pt tr:hover td{filter:brightness(.94);}'
        '#twrap:fullscreen{max-height:100vh;background:#fff;padding:14px;}'
        '#twrap:-webkit-full-screen{max-height:100vh;background:#fff;padding:14px;}'
        '</style>'
        '<script>'
        'var _sd={};'
        'function srt(c){'
        'var tb=document.getElementById("ptb");'
        'var rows=Array.from(tb.rows);'
        'var asc=_sd[c]!==true;_sd[c]=asc;'
        'rows.sort(function(a,b){'
        'var va=(a.cells[c]?a.cells[c].innerText:"").trim();'
        'var vb=(b.cells[c]?b.cells[c].innerText:"").trim();'
        'var na=parseFloat(va.replace(/[^0-9.]/g,""));'
        'var nb=parseFloat(vb.replace(/[^0-9.]/g,""));'
        'if(!isNaN(na)&&!isNaN(nb))return asc?na-nb:nb-na;'
        'return asc?va.localeCompare(vb,"fr"):vb.localeCompare(va,"fr");'
        '});'
        'rows.forEach(function(r){tb.appendChild(r);});'
        'for(var i=0;i<12;i++){'
        'var el=document.getElementById("si"+i);'
        'if(el)el.innerHTML=i===c?(asc?"&uarr;":"&darr;"):"&#8645;";'
        'if(el)el.style.opacity=i===c?"1":"0.45";'
        '}}'
        '</script>'
    )

    st.markdown("""
    <div class="legend-box">
        <div class="legend-item"><div class="legend-dot dot-green-dark"></div>Dernière activité ≤ 9 jours</div>
        <div class="legend-item"><div class="legend-dot dot-green-light"></div>Prochaine activité dans les 13 jours</div>
    </div>""", unsafe_allow_html=True)
    st.caption(f"{len(df_disp)} pistes affichées")
    import streamlit.components.v1 as components
    full_html = """<!DOCTYPE html>
<html><head>
<meta charset="utf-8">
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600&display=swap" rel="stylesheet">
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Space Grotesk', sans-serif; font-size: 0.79rem; background: white; }
</style>
</head><body>""" + table + "</body></html>"
    components.html(full_html, height=734, scrolling=True)
    # ── Export XLSX propre
    output = io.BytesIO()
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb_out = openpyxl.Workbook()
        ws = wb_out.active
        ws.title = "Pipeline"

        export_cols = list(df_disp.columns)

        # Styles
        fill_header = PatternFill("solid", fgColor="0A3D0A")
        fill_dark   = PatternFill("solid", fgColor="1A6E3C")
        fill_light  = PatternFill("solid", fgColor="6FCF97")
        fill_even   = PatternFill("solid", fgColor="F4FAF4")
        font_white  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
        font_normal = Font(size=9, name="Calibri")
        font_dark   = Font(color="0A3D0A", size=9, name="Calibri")
        align_wrap  = Alignment(wrap_text=True, vertical="top")
        align_mid   = Alignment(wrap_text=False, vertical="center")
        thin        = Side(style="thin", color="C8E6C8")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_widths = {
            "Alerte":20, "Fonds":22, "Typologie":20, "Entité":24,
            "Contact":18, "Contact Levée":18, "Dernier commentaire":60,
            "Intérêt":10, "Dernière activité":16, "Prochaine activité":16,
            "Ticket":12, "Matu.":10, "Tri":22,
        }

        # En-tête
        for ci, col in enumerate(export_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill      = fill_header
            cell.font      = font_white
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 16)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        now_ts2          = pd.Timestamp.now()
        cutoff_d2        = now_ts2 - timedelta(days=9)
        cutoff_p2        = now_ts2 + timedelta(days=13)

        # Données
        for ri, (_, row) in enumerate(df_disp.iterrows(), 2):
            date_d2 = pd.to_datetime(row.get("Dernière activité",""), dayfirst=True, errors="coerce")
            date_p2 = pd.to_datetime(row.get("Prochaine activité",""), dayfirst=True, errors="coerce")
            flag_d2 = pd.notna(date_d2) and date_d2 >= cutoff_d2
            flag_p2 = pd.notna(date_p2) and now_ts2 <= date_p2 <= cutoff_p2

            for ci, col in enumerate(export_cols, 1):
                val  = row.get(col, "")
                val  = "" if str(val).lower() in ["nan","none"] else val
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border

                if col == "Dernier commentaire":
                    cell.alignment = align_wrap
                else:
                    cell.alignment = align_mid

                if flag_d2 and col in ["Dernière activité","Tri","Dernier commentaire"]:
                    cell.fill = fill_dark
                    cell.font = Font(color="FFFFFF", size=9, name="Calibri")
                elif flag_p2 and col in ["Prochaine activité","Tri","Dernier commentaire"]:
                    cell.fill = fill_light
                    cell.font = font_dark
                elif ri % 2 == 0:
                    cell.fill = fill_even
                    cell.font = font_normal
                else:
                    cell.font = font_normal

            ws.row_dimensions[ri].height = 40

        wb_out.save(output)
    except Exception as e:
        df_disp.to_excel(output, index=False, sheet_name="Pipeline")

    output.seek(0)
    st.download_button(
        "⬇️ Exporter en Excel (.xlsx)",
        data=output,
        file_name=f"pipeline_exergon_{today.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
